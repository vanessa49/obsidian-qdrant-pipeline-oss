"""
访谈项目批量入库处理模块
支持三种处理路径：
- 路径A：有关键点Excel（优先）
- 路径B：有大纲但没关键点（次优）
- 路径C：只有转录稿
"""

import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from llm_client import call_llm


def scan_project(folder: str) -> dict:
    """
    扫描项目文件夹，识别文件角色
    
    Args:
        folder: 项目文件夹路径
        
    Returns:
        {
            "guide": "大纲文件路径或None",
            "keypoints": "关键点文件路径或None", 
            "transcripts": ["txt文件路径列表"]
        }
    """
    folder_path = Path(folder)
    
    if not folder_path.exists():
        raise ValueError(f"文件夹不存在: {folder}")
    
    if not folder_path.is_dir():
        raise ValueError(f"不是文件夹: {folder}")
    
    result = {
        "guide": None,
        "keypoints": None,
        "transcripts": []
    }
    
    # 扫描所有文件
    for file_path in folder_path.iterdir():
        if not file_path.is_file():
            continue
        
        filename_lower = file_path.stem.lower()
        suffix = file_path.suffix.lower()
        
        # 识别大纲文件
        if any(kw in filename_lower for kw in ["大纲", "guide", "outline",'提纲']):
            if suffix in [".xlsx", ".docx"]:
                result["guide"] = str(file_path)
                print(f"  识别为大纲: {file_path.name}")
            else:
                print(f"  警告: 文件名含'大纲'但格式不支持，忽略: {file_path.name}")
        
        # 识别关键点文件
        elif any(kw in filename_lower for kw in ["关键点", "keypoint", "key_point", "summary",'记录']):
            if suffix == ".xlsx":
                result["keypoints"] = str(file_path)
                print(f"  识别为关键点: {file_path.name}")
            else:
                print(f"  警告: 文件名含'关键点'但格式不是xlsx，忽略: {file_path.name}")
        
        # 识别转录稿
        elif suffix == ".txt":
            result["transcripts"].append(str(file_path))
            print(f"  识别为转录稿: {file_path.name}")
        
        # 其他格式忽略
        elif suffix in [".xlsx", ".docx"]:
            print(f"  警告: 未识别的文件，忽略: {file_path.name}")
    
    return result


def process_keypoints(keypoints_path: str, project_name: str) -> List[Dict]:
    """
    路径A：处理关键点Excel文件
    
    Args:
        keypoints_path: 关键点文件路径
        project_name: 项目名称
        
    Returns:
        chunk列表，每个chunk包含text和metadata
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")
    
    print(f"\n处理关键点文件: {Path(keypoints_path).name}")
    
    # 直接用 openpyxl 读取 Excel
    wb = openpyxl.load_workbook(keypoints_path, read_only=True)
    ws = wb.active
    
    # 读取所有数据
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    wb.close()
    
    if len(data) < 2:
        print(f"  ✗ Excel数据不足")
        return []
    
    # 判断表格方向：分析表格结构
    first_row = data[0]
    
    # 统计所有行的非空列数
    col_non_empty_counts = [0] * len(first_row)
    for row in data[:min(20, len(data))]:  # 检查前20行
        for col_idx, cell in enumerate(row):
            if cell and str(cell).strip():
                col_non_empty_counts[col_idx] += 1
    
    # 检查第一列的平均内容长度
    first_col_lengths = []
    for row in data[:min(20, len(data))]:
        if row and row[0]:
            first_col_lengths.append(len(str(row[0]).strip()))
    
    avg_first_col_length = sum(first_col_lengths) / len(first_col_lengths) if first_col_lengths else 0
    
    # 判断逻辑：
    # 1. 如果第一列平均长度>15（像问题），且其他列也有数据 → 横向表格
    # 2. 如果只有前2列有数据，但第一列很长 → 可能是纵向表格
    is_horizontal = False
    
    if avg_first_col_length > 15 and len([c for c in col_non_empty_counts[1:] if c > 0]) >= 1:
        # 第一列是长文本（问题），且至少有1列有回答 → 横向表格
        is_horizontal = True
        print(f"  → 第一列平均长度: {avg_first_col_length:.1f}, 识别为横向表格")
    elif len([c for c in col_non_empty_counts if c > 0]) == 2:
        # 只有2列有数据 → 纵向表格
        is_horizontal = False
        print(f"  → 只有2列有数据，识别为纵向表格")
    
    chunks = []
    
    if is_horizontal:
        # 横向表格：每列是一个受访者
        print(f"  → 识别为横向表格（每列=受访者）")
        
        # 第一列是问题，其余列是受访者回答
        # 获取受访者数量（有数据的列数-1）
        num_cols = len(first_row)
        
        # 为每个受访者列生成文档
        for col_idx in range(1, num_cols):
            # 收集该列的所有非空回答，判断是否真的有数据
            has_data = False
            for row in data:
                if col_idx < len(row) and row[col_idx] and str(row[col_idx]).strip():
                    has_data = True
                    break
            
            if not has_data:
                continue
            
            # 受访者名称（尝试从前几行提取）
            interviewee_name = None
            for row_idx in range(min(5, len(data))):
                if data[row_idx] and col_idx < len(data[row_idx]) and data[row_idx][col_idx]:
                    cell_value = str(data[row_idx][col_idx]).strip()
                    # 如果是简短的名字或标识（不是长段回答）
                    if cell_value and len(cell_value) < 30 and not any(c in cell_value for c in ['、', '。', '\n\n']):
                        interviewee_name = cell_value
                        break
            
            if not interviewee_name:
                interviewee_name = f"受访者{col_idx}"
            
            # 收集该受访者的所有问答
            qa_pairs = []
            for row_idx, row in enumerate(data):
                if not row or not row[0]:
                    continue
                
                question = str(row[0]).strip()
                answer = ""
                
                if col_idx < len(row) and row[col_idx]:
                    answer = str(row[col_idx]).strip()
                
                # 跳过没有回答的问题
                if not answer or answer in ['None', 'nan', 'NaN']:
                    continue
                
                # 跳过元信息行（年龄、时间等）
                if len(question) < 5:
                    continue
                
                qa_pairs.append(f"**{question}**\n{answer}")
            
            if not qa_pairs:
                continue
            
            # 合并为一个完整文档
            full_text = f"# {interviewee_name} - 访谈记录\n\n" + "\n\n".join(qa_pairs)
            
            chunks.append({
                "text": full_text,
                "metadata": {
                    "project": project_name,
                    "interviewee": interviewee_name,
                    "source_type": "interview_keypoints",
                    "tier": "processed",
                    "qa_count": len(qa_pairs)
                }
            })
        
        print(f"  ✓ 生成 {len(chunks)} 个受访者文档")
    
    else:
        # 纵向表格：每行是一个问答对
        print(f"  → 识别为纵向表格（每行=问答对）")
        
        for idx, row in enumerate(data[1:]):  # 跳过表头
            if not row or len(row) < 2:
                continue
            
            question = row[0] if row[0] else ""
            answer = row[1] if len(row) > 1 and row[1] else ""
            
            if not question or not answer:
                continue
            
            chunk_text = f"问题：{question}\n回答要点：{answer}"
            
            chunks.append({
                "text": chunk_text,
                "metadata": {
                    "project": project_name,
                    "source_type": "interview_keypoints",
                    "tier": "processed",
                    "chunk_index": idx
                }
            })
        
        print(f"  ✓ 提取 {len(chunks)} 个问答对")
    
    return chunks


def process_with_guide(transcript_path: str, guide_path: str, project_name: str) -> List[Dict]:
    """
    路径B：使用大纲文件辅助处理转录稿
    
    Args:
        transcript_path: 转录稿路径
        guide_path: 大纲文件路径
        project_name: 项目名称
        
    Returns:
        chunk列表（每个转录稿一个完整文档）
    """
    print(f"\n处理转录稿（使用大纲）: {Path(transcript_path).name}")
    
    # 从文件名提取受访者名称
    transcript_filename = Path(transcript_path).stem
    interviewee_name = transcript_filename.replace(project_name, "").strip("-_")
    if not interviewee_name:
        interviewee_name = transcript_filename
    
    # 1. 提取大纲问题列表
    questions = _extract_questions_from_guide(guide_path)
    
    if not questions:
        print(f"  警告: 未能从大纲提取问题，回退到无大纲模式")
        return process_transcript_only(transcript_path, project_name)
    
    print(f"  ✓ 从大纲提取 {len(questions)} 个问题")
    
    # 2. 读取转录稿
    transcript_text = Path(transcript_path).read_text(encoding='utf-8')
    
    # 截取前4000字符送给LLM
    transcript_sample = transcript_text[:4000]
    
    # 3. 构建问题列表字符串
    questions_str = "\n".join([f"{i+1}. {q}" for i, q in enumerate(questions)])
    
    # 4. 调用LLM提取问答对
    system_prompt = """你是质性研究助手。
只返回问答对内容，不要任何额外说明。"""
    
    user_prompt = f"""以下是访谈大纲的问题列表：

{questions_str}

以下是访谈转录原文（可能有语音识别错误，请自行判断）：

{transcript_sample}

请按照大纲问题，从转录稿中找出每个问题对应的回答。

对每个问题输出：

问题：[问题原文]
受访者回答：[从转录稿提取的相关内容，保留原话，修正明显的语音识别错误]

如果转录稿中没有该问题的明确回答，输出"未涉及"。

每个问答对之间用空行分隔。"""
    
    try:
        print(f"  → 调用LLM提取问答对...")
        llm_output = call_llm(prompt=user_prompt, system=system_prompt)
        
        if not llm_output:
            raise ValueError("LLM返回空内容")
        
        print(f"  ✓ LLM处理完成")
        
        # 5. 解析LLM输出为问答对列表
        qa_pairs = _parse_qa_pairs_text(llm_output)
        
        if not qa_pairs:
            print(f"  警告: 未能解析LLM输出，回退到直接切块")
            return _fallback_chunk(transcript_text, project_name, "interview_guided")
        
        print(f"  ✓ 提取 {len(qa_pairs)} 个问答对")
        
        # 6. 合并为一个完整文档（类似横向表格的处理方式）
        formatted_qa = [f"**{qa}**" if qa.startswith("问题：") else qa for qa in qa_pairs]
        full_text = f"# {interviewee_name} - 访谈记录\n\n" + "\n\n".join(formatted_qa)
        
        return [{
            "text": full_text,
            "metadata": {
                "project": project_name,
                "interviewee": interviewee_name,
                "source_type": "interview_guided",
                "tier": "processed",
                "qa_count": len(qa_pairs)
            }
        }]
    
    except Exception as e:
        print(f"  ✗ LLM处理失败: {e}")
        print(f"  → 回退到直接切块")
        return _fallback_chunk(transcript_text, project_name, "interview_guided")


def process_transcript_only(transcript_path: str, project_name: str) -> List[Dict]:
    """
    路径C：只有转录稿，直接切块
    
    Args:
        transcript_path: 转录稿路径
        project_name: 项目名称
        
    Returns:
        chunk列表（单个完整文档）
    """
    print(f"\n处理转录稿（无大纲）: {Path(transcript_path).name}")
    
    # 从文件名提取受访者名称
    transcript_filename = Path(transcript_path).stem
    interviewee_name = transcript_filename.replace(project_name, "").strip("-_")
    if not interviewee_name:
        interviewee_name = transcript_filename
    
    transcript_text = Path(transcript_path).read_text(encoding='utf-8')
    
    # 添加标题
    full_text = f"# {interviewee_name} - 访谈转录\n\n{transcript_text}"
    
    return [{
        "text": full_text,
        "metadata": {
            "project": project_name,
            "interviewee": interviewee_name,
            "source_type": "interview_raw",
            "tier": "raw_material",
            "original_length": len(transcript_text)
        }
    }]


def _extract_questions_from_guide(guide_path: str) -> List[str]:
    """
    从大纲文件提取问题列表
    
    Args:
        guide_path: 大纲文件路径
        
    Returns:
        问题列表
    """
    guide_path_obj = Path(guide_path)
    suffix = guide_path_obj.suffix.lower()
    
    questions = []
    
    try:
        if suffix == ".xlsx":
            # Excel文件：读取第一列或最长的一列
            try:
                import openpyxl
            except ImportError:
                print("  警告: 需要安装 openpyxl: pip install openpyxl")
                return []
            
            wb = openpyxl.load_workbook(guide_path, read_only=True)
            ws = wb.active
            
            # 读取第一列的所有非空单元格
            for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
                if row[0] and isinstance(row[0], str):
                    text = row[0].strip()
                    if text and len(text) > 3:  # 过滤太短的内容
                        questions.append(text)
            
            wb.close()
        
        elif suffix == ".docx":
            # Word文件：提取每行非空文字
            try:
                from markitdown import MarkItDown
            except ImportError:
                print("  警告: 需要安装 markitdown: pip install markitdown")
                return []
            
            md = MarkItDown()
            result = md.convert(guide_path)
            text = result.text_content
            
            # 按行分割，提取非空行
            for line in text.split('\n'):
                line = line.strip()
                # 移除Markdown标记
                line = re.sub(r'^#+\s*', '', line)
                line = re.sub(r'^\*+\s*', '', line)
                line = re.sub(r'^\d+\.\s*', '', line)
                
                if line and len(line) > 3:
                    questions.append(line)
        
    except Exception as e:
        print(f"  警告: 提取问题失败: {e}")
    
    return questions


def _parse_qa_pairs(llm_output: str, project_name: str) -> List[Dict]:
    """
    解析LLM输出的问答对（旧版本，保留用于兼容）
    
    Args:
        llm_output: LLM输出文本
        project_name: 项目名称
        
    Returns:
        chunk列表
    """
    chunks = []
    
    # 按空行分割
    blocks = re.split(r'\n\s*\n', llm_output)
    
    for idx, block in enumerate(blocks):
        block = block.strip()
        if not block:
            continue
        
        # 检查是否包含"问题："和"回答"
        if '问题：' in block or '问题:' in block:
            chunks.append({
                "text": block,
                "metadata": {
                    "project": project_name,
                    "source_type": "interview_guided",
                    "tier": "processed",
                    "chunk_index": idx
                }
            })
    
    return chunks


def _parse_qa_pairs_text(llm_output: str) -> List[str]:
    """
    解析LLM输出的问答对为文本列表
    
    Args:
        llm_output: LLM输出文本
        
    Returns:
        问答对文本列表
    """
    qa_pairs = []
    
    # 按空行分割
    blocks = re.split(r'\n\s*\n', llm_output)
    
    for block in blocks:
        block = block.strip()
        if not block:
            continue
        
        # 检查是否包含"问题："
        if '问题：' in block or '问题:' in block:
            qa_pairs.append(block)
    
    return qa_pairs


def _fallback_chunk(text: str, project_name: str, source_type: str) -> List[Dict]:
    """
    回退策略：直接按固定大小切块
    
    Args:
        text: 文本内容
        project_name: 项目名称
        source_type: 来源类型
        
    Returns:
        chunk列表
    """
    CHUNK_SIZE = 800
    chunks = []
    
    # 按段落分割
    paragraphs = re.split(r'\n\n+', text)
    
    current_chunk = ""
    chunk_index = 0
    
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        
        if len(current_chunk) + len(para) + 2 <= CHUNK_SIZE:
            if current_chunk:
                current_chunk += "\n\n" + para
            else:
                current_chunk = para
        else:
            # 保存当前chunk
            if current_chunk:
                chunks.append({
                    "text": current_chunk,
                    "metadata": {
                        "project": project_name,
                        "source_type": source_type,
                        "tier": "raw_material",
                        "chunk_index": chunk_index
                    }
                })
                chunk_index += 1
            
            current_chunk = para
    
    # 添加最后一个chunk
    if current_chunk:
        chunks.append({
            "text": current_chunk,
            "metadata": {
                "project": project_name,
                "source_type": source_type,
                "tier": "raw_material",
                "chunk_index": chunk_index
            }
        })
    
    print(f"  ✓ 切分为 {len(chunks)} 个chunks")
    
    return chunks


def determine_processing_path(scan_result: dict) -> Tuple[str, str]:
    """
    根据扫描结果确定处理路径
    
    Args:
        scan_result: scan_project的返回结果
        
    Returns:
        (路径名称, 路径描述)
    """
    if scan_result["keypoints"]:
        return "A", "关键点优先（直接结构化）"
    elif scan_result["guide"]:
        return "B", "大纲辅助（LLM提取问答对）"
    else:
        return "C", "仅转录稿（直接切块）"


if __name__ == "__main__":
    # 测试代码
    import sys
    
    if len(sys.argv) < 2:
        print("用法: python interview_processor.py <项目文件夹路径>")
        sys.exit(1)
    
    folder = sys.argv[1]
    
    print("="*60)
    print("访谈项目扫描测试")
    print("="*60)
    
    try:
        result = scan_project(folder)
        
        print("\n扫描结果:")
        print(f"  大纲: {result['guide'] or '无'}")
        print(f"  关键点: {result['keypoints'] or '无'}")
        print(f"  转录稿: {len(result['transcripts'])} 个")
        
        path, desc = determine_processing_path(result)
        print(f"\n将使用路径{path}: {desc}")
        
    except Exception as e:
        print(f"✗ 错误: {e}")
